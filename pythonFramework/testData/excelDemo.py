import openpyxl

# Load the Excel Sheet
book = openpyxl.load_workbook("C:\\Users\\PMLS\\Desktop\\pythonExcelDemo\\PythonDemo.xlsx")
sheet = book.active
# To read data from excel sheet specific cell provide the cell address in the form of row and column
cell = sheet.cell(row=1, column=2)
print(cell.value)
# Write data in the excel sheet cell
sheet.cell(row=2, column=2).value = "Zuni"
print(sheet.cell(row=2, column=2).value)
# To get the maximum rows in the excel sheet
print(sheet.max_row)
# To get the maximum column in the excel sheet
print(sheet.max_column)
# another method to read the value from the Excel sheet cell
print(sheet['C1'].value)
# To print every value present in the excel sheet
# Initializing the empty dictionary and storing excel data into it.
Dict = {}
# Syntax to store data into Dictionary    Dict["firstName"] = "Zuni"
for i in range(1, sheet.max_row+1): # if there are 1-5 rows then range will run the loop runs only 1-4 thats why we put +1 so that we can get all the rows from the excel sheet.
    if sheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)