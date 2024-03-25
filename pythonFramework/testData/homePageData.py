import openpyxl
class homePageDataset:
    test_HomePage = [{"firstName": "Zuni", "email": "zuni@gmail.com", "password": "123", "gender": "Female"}, {"firstName": "Afaq", "email": "afaq@gmail.com", "password": "456", "gender": "Male"}]

    @staticmethod
    def getTestData(testCaseName):
        Dict = {}
        # open Excel sheet
        book = openpyxl.load_workbook("C:\\Users\\PMLS\\Desktop\\pythonExcelDemo\\PythonDemo.xlsx")
        sheet = book.active
        # Syntax to store data into Dictionary    Dict["firstName"] = "Zuni"
        for i in range(1, sheet.max_row + 1):  # if there are 1-5 rows then range will run the loop runs only 1-4 thats why we put +1 so that we can get all the rows from the excel sheet.
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

