from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl


def updateExcelData(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(filePath)

fruitName = "Apple"
newValue = "999"
filePath = "C:\\Users\\PMLS\\Downloads\\download.xlsx"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
# Click on download btn
driver.find_element(By.ID, "downloadButton").click()
# update the downloaded Excel file
updateExcelData(filePath, fruitName, "price", newValue)
# Upload the updated file
fileInput = driver.find_element(By.CSS_SELECTOR, "input[type = 'file']")
fileInput.send_keys(filePath)
# Wait until the toast msg appears
wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")))
# Creating a dynamic Xpath which can use for every value in the table
# Generating Xpath to get the actual price from the table

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actualPrice)
assert actualPrice == newValue
